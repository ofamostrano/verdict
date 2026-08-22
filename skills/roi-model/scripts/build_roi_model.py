#!/usr/bin/env python3
"""Generate an investor-grade ROI workbook with live formulas.

Produces eight tabs: README, Assumptions, Unit Economics, AI Cost, P&L, Cash,
Sensitivity, Scenarios. All computed cells reference named ranges on the
Assumptions tab, so a reviewer can change one input and watch every dependent
figure move.

Usage:
    python build_roi_model.py --template > assumptions.json
    python build_roi_model.py assumptions.json --out model.xlsx

Assumption grades: GREEN (sourced), AMBER (founder estimate), RED (placeholder).
Every assumption row carries a grade and a source string.
"""

import argparse
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

MONTHS = 36

GRADE_FILLS = {
    "GREEN": PatternFill("solid", fgColor="C6EFCE"),
    "AMBER": PatternFill("solid", fgColor="FFEB9C"),
    "RED": PatternFill("solid", fgColor="FFC7CE"),
}
HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14, color="1F3864")
SECTION_FONT = Font(bold=True, size=11, color="1F3864")
MONEY = '#,##0'
MONEY2 = '#,##0.00'
PCT = '0.0%'
RATIO = '0.00'
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# name, label, value, format, grade, source
DEFAULT_ASSUMPTIONS = [
    ["company_name", "Company name", "Venture", "text", "GREEN", "Founder"],
    ["price_monthly", "Price per account per month ($)", 299, MONEY, "AMBER", "Comparable pricing; untested"],
    ["new_accounts_m1", "New accounts in month 1", 6, "0", "AMBER", "Founder estimate"],
    ["growth_new_accts", "Monthly growth in new accounts", 0.18, PCT, "AMBER", "Founder estimate"],
    ["churn_monthly", "Monthly logo churn", 0.03, PCT, "AMBER", "SMB SaaS norm; untested"],
    ["expansion_monthly", "Monthly revenue expansion per account", 0.01, PCT, "AMBER", "Founder estimate"],
    ["cac", "Fully loaded CAC ($)", 1200, MONEY, "AMBER", "Founder estimate incl. loaded salary"],
    ["hosting_per_acct", "Hosting and infrastructure per account/mo ($)", 12, MONEY2, "AMBER", "Cloud calculator"],
    ["payment_fee_pct", "Payment processing fee", 0.029, PCT, "GREEN", "Processor published rate"],
    ["tasks_per_acct", "AI tasks per account per month", 120, "0", "AMBER", "Workflow estimate"],
    ["cost_per_call", "Inference + tool cost per attempt ($)", 0.035, MONEY2, "AMBER", "Model pricing at est. token use"],
    ["task_success_rate", "Task success rate (no human correction)", 0.88, PCT, "AMBER", "Target; measure post-launch"],
    ["retries_per_task", "Average attempts per completed task", 1.25, RATIO, "AMBER", "Founder estimate"],
    ["intervention_rate", "Human intervention rate", 0.12, PCT, "AMBER", "Target; measure post-launch"],
    ["intervention_cost", "Fully loaded cost per intervention ($)", 4.50, MONEY2, "AMBER", "Loaded hourly / tasks per hour"],
    ["opex_fixed", "Fixed monthly operating expense ($)", 14000, MONEY, "AMBER", "Founder + contractors + tooling"],
    ["opex_growth", "Monthly growth in fixed opex", 0.02, PCT, "AMBER", "Founder estimate"],
    ["sm_spend_pct", "Sales and marketing as share of revenue", 0.25, PCT, "AMBER", "Founder estimate"],
    ["opening_cash", "Opening cash ($)", 150000, MONEY, "GREEN", "Bank balance / committed raise"],
]

SENSITIVITY_VARS = [
    ("churn_monthly", "Monthly churn"),
    ("cac", "CAC"),
    ("price_monthly", "Price"),
    ("cost_per_call", "AI cost per attempt"),
    ("task_success_rate", "Task success rate"),
    ("intervention_rate", "Human intervention rate"),
]


def style_header(ws, row, last_col):
    for col in range(1, last_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BOX


def set_widths(ws, widths):
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def build_readme(wb, assumptions):
    ws = wb.create_sheet("README")
    company = next((a[2] for a in assumptions if a[0] == "company_name"), "Venture")
    ws["A1"] = f"{company} — ROI and Unit Economics Model"
    ws["A1"].font = TITLE_FONT
    rows = [
        "",
        ["Purpose", "Investor-grade model: buyer-side ROI, unit economics, 36-month P&L and cash, sensitivity, scenarios."],
        ["How to use", "Change any cell on the Assumptions tab. Every other tab recalculates. No hard-coded values live outside Assumptions."],
        "",
        ["Assumption grading", ""],
        ["GREEN", "Sourced from a named external source or the venture's own measured data."],
        ["AMBER", "Founder estimate: reasonable but untested. Each needs a named cheapest test."],
        ["RED", "Placeholder. Must be replaced before this model is shown to anyone."],
        "",
        ["Definitions", ""],
        ["LTV", "ARPA x gross margin / monthly churn. Computed from gross profit, never from revenue."],
        ["CAC payback", "CAC / monthly gross profit per account, in months."],
        ["Cost per successful task", "Total inference and tool spend / tasks completed without human correction, plus fallback labor."],
        ["Burn multiple", "Net burn / net new ARR."],
        "",
        ["Thresholds", ""],
        ["LTV/CAC", ">= 3.0"],
        ["CAC payback", "<= 18 months (<= 12 for SMB motions)"],
        ["Gross margin", ">= 70% for software"],
        ["Seed evidence bar", "~$10k MRR or ~1,000 engaged users"],
    ]
    r = 3
    for row in rows:
        if row == "":
            r += 1
            continue
        ws.cell(row=r, column=1, value=row[0]).font = SECTION_FONT
        ws.cell(row=r, column=2, value=row[1]).alignment = Alignment(wrap_text=True, vertical="top")
        r += 1
    set_widths(ws, [26, 96])
    for grade, fill in GRADE_FILLS.items():
        for row_idx in range(3, r):
            if ws.cell(row=row_idx, column=1).value == grade:
                ws.cell(row=row_idx, column=1).fill = fill
    return ws


def build_assumptions(wb, assumptions):
    ws = wb.create_sheet("Assumptions")
    ws["A1"] = "Assumptions — single source of truth"
    ws["A1"].font = TITLE_FONT
    headers = ["Name", "Assumption", "Value", "Grade", "Source / cheapest test to upgrade"]
    for col, head in enumerate(headers, start=1):
        ws.cell(row=3, column=col, value=head)
    style_header(ws, 3, len(headers))

    for offset, (name, label, value, fmt, grade, source) in enumerate(assumptions):
        row = 4 + offset
        ws.cell(row=row, column=1, value=name).font = Font(italic=True, size=9, color="808080")
        ws.cell(row=row, column=2, value=label)
        cell = ws.cell(row=row, column=3, value=value)
        if fmt != "text":
            cell.number_format = fmt
        cell.font = Font(bold=True)
        grade_cell = ws.cell(row=row, column=4, value=grade)
        grade_cell.fill = GRADE_FILLS.get(grade, GRADE_FILLS["AMBER"])
        grade_cell.alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=5, value=source).alignment = Alignment(wrap_text=True)
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = BOX
        wb.defined_names.add(DefinedName(name, attr_text=f"Assumptions!$C${row}"))

    set_widths(ws, [22, 44, 16, 10, 52])
    ws.freeze_panes = "A4"
    return ws


def build_unit_economics(wb):
    ws = wb.create_sheet("Unit Economics")
    ws["A1"] = "Unit Economics"
    ws["A1"].font = TITLE_FONT
    headers = ["Metric", "Value", "Threshold", "Verdict", "Note"]
    for col, head in enumerate(headers, start=1):
        ws.cell(row=3, column=col, value=head)
    style_header(ws, 3, len(headers))

    rows = [
        ("ARPA — revenue per account per month", "=price_monthly", MONEY, "", "", "From Assumptions"),
        ("Payment processing cost", "=price_monthly*payment_fee_pct", MONEY2, "", "", "Percent of revenue"),
        ("Hosting cost per account", "=hosting_per_acct", MONEY2, "", "", ""),
        ("AI delivered cost per account", "='AI Cost'!B12", MONEY2, "", "", "Includes retries and fallback labor"),
        ("Total COGS per account", "=B5+B6+B7", MONEY2, "", "", "Payments + hosting + AI delivered"),
        ("Gross profit per account", "=B4-B8", MONEY2, "", "", ""),
        ("Gross margin", "=IF(B4=0,0,B9/B4)", PCT, 0.70, '=IF(B10>=C10,"PASS","REVIEW")', "AI cost and fallback labor break this"),
        ("Monthly logo churn", "=churn_monthly", PCT, 0.03, '=IF(B11<=C11,"PASS","REVIEW")', "Dominates LTV"),
        ("Average account lifetime (months)", "=IF(churn_monthly=0,0,1/churn_monthly)", RATIO, "", "", ""),
        ("LTV (gross profit basis)", "=B9*B12", MONEY, "", "", "Never computed from revenue"),
        ("CAC (fully loaded)", "=cac", MONEY, "", "", "Includes loaded salary"),
        ("LTV / CAC", "=IF(cac=0,0,B13/B14)", RATIO, 3.0, '=IF(B15>=C15,"PASS","REVIEW")', "Below 3 needs a named reason"),
        ("CAC payback (months)", "=IF(B9=0,0,B14/B9)", RATIO, 18, '=IF(B16<=C16,"PASS","REVIEW")', "Gross profit basis, not revenue"),
        ("Accounts needed for $10k MRR", "=IF(price_monthly=0,0,10000/price_monthly)", RATIO, "", "", "The seed evidence bar in accounts"),
        ("MRR at final month", f"='P&L'!{get_column_letter(1 + MONTHS)}6", MONEY, "", "", "From the P&L tab"),
    ]
    for offset, (label, formula, fmt, threshold, verdict, note) in enumerate(rows):
        row = 4 + offset
        ws.cell(row=row, column=1, value=label)
        cell = ws.cell(row=row, column=2, value=formula)
        cell.number_format = fmt
        if threshold != "":
            tc = ws.cell(row=row, column=3, value=threshold)
            tc.number_format = fmt
        if verdict:
            vc = ws.cell(row=row, column=4, value=verdict)
            vc.alignment = Alignment(horizontal="center")
            vc.font = Font(bold=True)
        ws.cell(row=row, column=5, value=note).alignment = Alignment(wrap_text=True)
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = BOX

    set_widths(ws, [42, 16, 14, 12, 48])
    return ws


def build_ai_cost(wb):
    ws = wb.create_sheet("AI Cost")
    ws["A1"] = "AI Cost — per successful task, not per call"
    ws["A1"].font = TITLE_FONT
    headers = ["Line", "Value", "Formula basis"]
    for col, head in enumerate(headers, start=1):
        ws.cell(row=3, column=col, value=head)
    style_header(ws, 3, len(headers))

    rows = [
        ("Tasks attempted per account per month", "=tasks_per_acct*retries_per_task", RATIO, "Volume x average attempts per completion"),
        ("Tasks completed per account per month", "=tasks_per_acct", RATIO, "From Assumptions"),
        ("Tasks succeeding without human correction", "=tasks_per_acct*task_success_rate", RATIO, "Completion x success rate"),
        ("Inference and tool spend per account", "=B4*cost_per_call", MONEY2, "Attempts x cost per attempt"),
        ("Cost per attempted task", "=IF(B4=0,0,B7/B4)", MONEY2, "Naive figure — understates true cost"),
        ("Cost per successful task (inference only)", "=IF(B6=0,0,B7/B6)", MONEY2, "Spend / successful tasks"),
        ("Interventions per account per month", "=tasks_per_acct*intervention_rate", RATIO, "Volume x intervention rate"),
        ("Fallback labour cost per account", "=B10*intervention_cost", MONEY2, "Interventions x loaded cost each"),
        ("Delivered AI cost per account per month", "=B7+B11", MONEY2, "Inference + fallback labour"),
        ("Delivered cost per successful task", "=IF(B6=0,0,B12/B6)", MONEY2, "The number that belongs in the pitch"),
        ("AI cost as share of price", "=IF(price_monthly=0,0,B12/price_monthly)", PCT, "Gross margin pressure from AI"),
    ]
    for offset, (label, formula, fmt, basis) in enumerate(rows):
        row = 4 + offset
        ws.cell(row=row, column=1, value=label)
        cell = ws.cell(row=row, column=2, value=formula)
        cell.number_format = fmt
        ws.cell(row=row, column=3, value=basis).alignment = Alignment(wrap_text=True)
        for col in range(1, 4):
            ws.cell(row=row, column=col).border = BOX

    ws.cell(row=17, column=1, value="Why this tab exists").font = SECTION_FONT
    ws.cell(
        row=18, column=1,
        value=("An agent costing little per call but failing often is more expensive per completed task than one "
               "costing more per call and rarely failing. Reliability is therefore a gross-margin input, not only a "
               "quality metric, and the human-intervention rate is where a services business hides inside a software "
               "gross margin. Measure success rate and intervention rate in production and replace the targets here."),
    ).alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=18, start_column=1, end_row=21, end_column=3)
    set_widths(ws, [44, 16, 52])
    return ws


def build_pnl(wb):
    ws = wb.create_sheet("P&L")
    ws["A1"] = "P&L — 36 months (monthly granularity reveals the cash trough)"
    ws["A1"].font = TITLE_FONT

    ws.cell(row=3, column=1, value="Line")
    for m in range(1, MONTHS + 1):
        ws.cell(row=3, column=1 + m, value=f"M{m}")
    style_header(ws, 3, MONTHS + 1)

    labels = [
        "New accounts", "Active accounts", "MRR", "Revenue",
        "COGS", "Gross profit", "Gross margin",
        "Sales & marketing", "Fixed opex", "Total opex", "EBITDA",
    ]
    for offset, label in enumerate(labels):
        ws.cell(row=4 + offset, column=1, value=label).font = Font(bold=(label in ("Revenue", "Gross profit", "EBITDA")))

    for m in range(1, MONTHS + 1):
        col = get_column_letter(1 + m)
        prev = get_column_letter(m)
        # New accounts
        ws[f"{col}4"] = "=new_accounts_m1" if m == 1 else f"={prev}4*(1+growth_new_accts)"
        ws[f"{col}4"].number_format = RATIO
        # Active accounts
        ws[f"{col}5"] = f"={col}4" if m == 1 else f"={prev}5*(1-churn_monthly)+{col}4"
        ws[f"{col}5"].number_format = RATIO
        # MRR with expansion
        ws[f"{col}6"] = f"={col}5*price_monthly*(1+expansion_monthly)^{m - 1}"
        ws[f"{col}6"].number_format = MONEY
        # Revenue
        ws[f"{col}7"] = f"={col}6"
        ws[f"{col}7"].number_format = MONEY
        # COGS
        ws[f"{col}8"] = f"={col}5*('Unit Economics'!$B$8)"
        ws[f"{col}8"].number_format = MONEY
        # Gross profit
        ws[f"{col}9"] = f"={col}7-{col}8"
        ws[f"{col}9"].number_format = MONEY
        ws[f"{col}10"] = f"=IF({col}7=0,0,{col}9/{col}7)"
        ws[f"{col}10"].number_format = PCT
        # S&M: CAC-driven, floored by percent-of-revenue commitment
        ws[f"{col}11"] = f"=MAX({col}4*cac,{col}7*sm_spend_pct)"
        ws[f"{col}11"].number_format = MONEY
        ws[f"{col}12"] = "=opex_fixed" if m == 1 else f"={prev}12*(1+opex_growth)"
        ws[f"{col}12"].number_format = MONEY
        ws[f"{col}13"] = f"={col}11+{col}12"
        ws[f"{col}13"].number_format = MONEY
        ws[f"{col}14"] = f"={col}9-{col}13"
        ws[f"{col}14"].number_format = MONEY

    ws.freeze_panes = "B4"
    set_widths(ws, [24] + [11] * MONTHS)
    return ws


def build_cash(wb):
    ws = wb.create_sheet("Cash")
    ws["A1"] = "Cash and runway"
    ws["A1"].font = TITLE_FONT
    ws.cell(row=3, column=1, value="Line")
    for m in range(1, MONTHS + 1):
        ws.cell(row=3, column=1 + m, value=f"M{m}")
    style_header(ws, 3, MONTHS + 1)

    for offset, label in enumerate(["Opening cash", "EBITDA (net burn)", "Closing cash", "Cash negative?"]):
        ws.cell(row=4 + offset, column=1, value=label)

    for m in range(1, MONTHS + 1):
        col = get_column_letter(1 + m)
        prev = get_column_letter(m)
        ws[f"{col}4"] = "=opening_cash" if m == 1 else f"={prev}6"
        ws[f"{col}4"].number_format = MONEY
        ws[f"{col}5"] = f"='P&L'!{col}14"
        ws[f"{col}5"].number_format = MONEY
        ws[f"{col}6"] = f"={col}4+{col}5"
        ws[f"{col}6"].number_format = MONEY
        ws[f"{col}7"] = f'=IF({col}6<0,"YES","")'
        ws[f"{col}7"].alignment = Alignment(horizontal="center")

    summary_row = 10
    ws.cell(row=summary_row, column=1, value="Summary").font = SECTION_FONT
    last = get_column_letter(1 + MONTHS)
    m12 = get_column_letter(1 + min(12, MONTHS))
    items = [
        ("Minimum cash across 36 months", f"=MIN(B6:{last}6)", MONEY,
         "If negative, this is the size of the hole a raise must cover."),
        ("First month cash goes negative", f'=IFERROR(MATCH(TRUE,INDEX(B6:{last}6<0,0),0),"never")', "0",
         "Month index; 'never' means the plan is self-funding."),
        ("Runway at current burn (months)", f'=IF(B5>=0,"profitable",IFERROR(B4/-B5,"n/a"))', RATIO,
         "Opening cash divided by month-1 burn."),
        ("Month EBITDA first turns positive", f'=IFERROR(MATCH(TRUE,INDEX(\'P&L\'!B14:{last}14>0,0),0),"not in 36m")', "0",
         "The month the business stops consuming cash."),
        ("Net new ARR (months 1-12)", f"='P&L'!{m12}6*12", MONEY, "Annualised exit MRR of year one."),
        ("Burn multiple (year 1)", f"=IFERROR(-SUM(B5:{m12}5)/('P&L'!{m12}6*12),\"n/a\")", RATIO,
         "Net burn / net new ARR. Under 2.0 is efficient."),
    ]
    for offset, (label, formula, fmt, note) in enumerate(items):
        row = summary_row + 1 + offset
        ws.cell(row=row, column=1, value=label)
        cell = ws.cell(row=row, column=2, value=formula)
        cell.number_format = fmt
        cell.font = Font(bold=True)
        ws.cell(row=row, column=3, value=note).alignment = Alignment(wrap_text=True)

    ws.freeze_panes = "B4"
    set_widths(ws, [34, 14, 58] + [11] * (MONTHS - 2))
    return ws


def build_sensitivity(wb):
    ws = wb.create_sheet("Sensitivity")
    ws["A1"] = "Sensitivity — one variable at a time, ranked by swing"
    ws["A1"].font = TITLE_FONT
    ws.cell(row=2, column=1, value=(
        "Enter the low and high output values by changing each input on the Assumptions tab and reading the output "
        "metric. Rank by absolute swing: the top row is the number the business is actually betting on."
    )).alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=7)

    headers = ["Input", "Base value", "Low (-30%)", "High (+30%)",
               "Output at low", "Output at high", "Absolute swing"]
    for col, head in enumerate(headers, start=1):
        ws.cell(row=4, column=col, value=head)
    style_header(ws, 4, len(headers))

    ws.cell(row=5, column=1, value="Output metric tested:")
    ws.cell(row=5, column=2, value="Minimum cash (Cash!B11)").font = Font(bold=True)

    for offset, (name, label) in enumerate(SENSITIVITY_VARS):
        row = 7 + offset
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=f"={name}")
        ws.cell(row=row, column=3, value=f"={name}*0.7")
        ws.cell(row=row, column=4, value=f"={name}*1.3")
        ws.cell(row=row, column=5, value=None)
        ws.cell(row=row, column=6, value=None)
        ws.cell(row=row, column=7, value=f"=IFERROR(ABS(F{row}-E{row}),\"\")")
        ws.cell(row=row, column=7).number_format = MONEY
        for col in range(1, 8):
            ws.cell(row=row, column=col).border = BOX

    breakeven_row = 7 + len(SENSITIVITY_VARS) + 2
    ws.cell(row=breakeven_row, column=1, value="Breakevens — state these explicitly").font = SECTION_FONT
    for offset, (label, note) in enumerate([
        ("Churn at which LTV/CAC falls to 3.0", "Solve by changing churn_monthly until Unit Economics!B15 = 3."),
        ("CAC at which payback reaches 18 months", "Solve by changing cac until Unit Economics!B16 = 18."),
        ("Price at which gross margin falls to 70%", "Solve by changing price_monthly until Unit Economics!B10 = 70%."),
        ("Success rate at which AI cost consumes 30% of price", "Solve on the AI Cost tab, line 14."),
    ]):
        row = breakeven_row + 1 + offset
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value="")
        ws.cell(row=row, column=3, value=note).alignment = Alignment(wrap_text=True)

    chart = BarChart()
    chart.type = "bar"
    chart.title = "Tornado — absolute output swing by input"
    chart.y_axis.title = "Swing"
    data = Reference(ws, min_col=7, min_row=6, max_row=6 + len(SENSITIVITY_VARS))
    cats = Reference(ws, min_col=1, min_row=7, max_row=6 + len(SENSITIVITY_VARS))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height, chart.width = 8, 16
    ws.add_chart(chart, "I6")

    set_widths(ws, [38, 16, 16, 16, 16, 16, 16])
    return ws


def build_scenarios(wb):
    ws = wb.create_sheet("Scenarios")
    ws["A1"] = "Scenarios — every case names its trigger"
    ws["A1"].font = TITLE_FONT
    ws.cell(row=2, column=1, value=(
        "A scenario without a named trigger is an arithmetic exercise. Name the event that would cause each case "
        "and the leading indicator that would show it happening early enough to react."
    )).alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)

    headers = ["Scenario", "Trigger event", "Leading indicator",
               "Key input changes", "Minimum cash", "Month EBITDA positive"]
    for col, head in enumerate(headers, start=1):
        ws.cell(row=4, column=col, value=head)
    style_header(ws, 4, len(headers))

    rows = [
        ("Base", "Neither bear nor bull trigger fires; plan hiring against this case.",
         "Pipeline conversion holds within 20% of plan.", "As per Assumptions tab."),
        ("Bear", "A named competitor ships the wedge feature, or a regulatory deadline slips.",
         "Win rate in competitive deals declines two months running.",
         "Churn +50%, CAC +40%, new accounts -30%."),
        ("Bull", "A channel partnership lands, or a compliance deadline pulls demand forward.",
         "Inbound volume doubles month over month.",
         "New accounts +60%, churn -25%, price +15%."),
    ]
    for offset, (name, trigger, indicator, changes) in enumerate(rows):
        row = 5 + offset
        ws.cell(row=row, column=1, value=name).font = Font(bold=True)
        for col, value in enumerate([trigger, indicator, changes], start=2):
            ws.cell(row=row, column=col, value=value).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=row, column=5, value=None).number_format = MONEY
        ws.cell(row=row, column=6, value=None)
        for col in range(1, 7):
            ws.cell(row=row, column=col).border = BOX

    ws.cell(row=10, column=1, value="Method").font = SECTION_FONT
    ws.cell(row=11, column=1, value=(
        "Apply each scenario's input changes on the Assumptions tab, then record minimum cash (Cash!B11) and the "
        "month EBITDA turns positive (Cash!B14) here before restoring the base case. Record the ask against the "
        "bear case minimum cash, not the base case: raising against the optimistic path is how bridge rounds happen."
    )).alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=11, start_column=1, end_row=14, end_column=6)

    set_widths(ws, [12, 42, 34, 34, 16, 20])
    return ws


def main():
    parser = argparse.ArgumentParser(description="Generate an investor-grade ROI workbook with live formulas.")
    parser.add_argument("assumptions", nargs="?", help="Path to an assumptions JSON file.")
    parser.add_argument("--out", default="roi_model.xlsx", help="Output .xlsx path.")
    parser.add_argument("--template", action="store_true", help="Print a starting assumptions JSON and exit.")
    args = parser.parse_args()

    if args.template:
        print(json.dumps(
            [{"name": n, "label": l, "value": v, "format": f, "grade": g, "source": s}
             for n, l, v, f, g, s in DEFAULT_ASSUMPTIONS],
            indent=2,
        ))
        return

    assumptions = DEFAULT_ASSUMPTIONS
    if args.assumptions:
        raw = json.loads(Path(args.assumptions).read_text(encoding="utf-8"))
        assumptions = [
            [item["name"], item["label"], item["value"],
             item.get("format", MONEY), item.get("grade", "AMBER"), item.get("source", "")]
            for item in raw
        ]

    wb = Workbook()
    wb.remove(wb.active)
    build_readme(wb, assumptions)
    build_assumptions(wb, assumptions)
    build_unit_economics(wb)
    build_ai_cost(wb)
    build_pnl(wb)
    build_cash(wb)
    build_sensitivity(wb)
    build_scenarios(wb)
    wb.save(args.out)

    ambers = sum(1 for a in assumptions if a[4] == "AMBER")
    reds = sum(1 for a in assumptions if a[4] == "RED")
    print(f"Wrote {args.out} — 8 tabs, {len(assumptions)} assumptions ({ambers} AMBER, {reds} RED).")
    if reds:
        print("RED placeholders must be replaced before showing this model to anyone.")
    print("Next: populate the Sensitivity output columns and the Scenarios result columns by hand, "
          "then record breakevens.")


if __name__ == "__main__":
    main()
